from flask import Flask, request, send_file, redirect, render_template_string, url_for
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import xlsxwriter
import os
import tempfile

app = Flask(__name__)
app.secret_key = 'supersecretkey'

def create_table(ws, df, table_name):
    table = Table(displayName=table_name, ref=f"A1:{chr(64+len(df.columns))}{len(df)+1}")
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=True
    )
    table.tableStyleInfo = style
    ws.add_table(table)

# Function to create and insert a combined column and line chart
def create_combined_chart(workbook, worksheet, col_series_range, line_series_range, chart_position, chart_title):
    column_chart = workbook.add_chart({"type": "column"})
    column_chart.add_series({
        "values": col_series_range,
        "data_labels": {"value": True},
    })
    
    line_chart = workbook.add_chart({"type": "line"})
    line_chart.add_series({
        'categories': line_series_range[0],
        'values': line_series_range[1],
        "data_labels": {"value": True, 'position': 'above'},
    })
    
    column_chart.combine(line_chart)
    column_chart.set_title({"name": chart_title})
    column_chart.set_x_axis({"label_position": "low"})
    worksheet.insert_chart(chart_position, column_chart)

# Function to create and insert a column chart
def create_column_chart(workbook, worksheet, categories_range, values_range, chart_position, chart_title):
    column_chart = workbook.add_chart({"type": "column"})
    column_chart.add_series({
        "categories": categories_range,
        "values": values_range,
        "data_labels": {"value": True},
    })
    
    column_chart.set_title({"name": chart_title})
    column_chart.set_x_axis({"label_position": "low"})
    column_chart.set_legend({"none": True})
    worksheet.insert_chart(chart_position, column_chart, {"x_offset": 25, "y_offset": 10})

def create_custom_chart(workbook, worksheet, chart_position):
    chart = workbook.add_chart({"type": "column"})
    
    chart.add_series({
        "name": "=強積⾦表現!$U$1",
        "categories": "=強積⾦表現!$A$6:$A$14",
        "values": "=強積⾦表現!$U$6:$U$14",
        "fill": {"color": "yellow"}, 
        "data_labels": {"value": True, "border": {}, "fill": {}, "font": {"color": "green"}},
    })
    
    chart.add_series({
        "name": ["強積⾦表現", 0, 21],
        "categories": ["強積⾦表現", 5, 0, 13, 0],
        "values": ["強積⾦表現", 5, 21, 13, 21],
        "fill": {"color": "yellow"}, 
        "data_labels": {"value": True, "border": {}, "fill": {}, "font": {"color": "red"}},
    })
    
    chart.add_series({
        "name": ["強積⾦表現", 0, 22],
        "categories": ["強積⾦表現", 5, 0, 13, 0],
        "values": ["強積⾦表現", 5, 22, 13, 22],
        "fill": {"color": "black"}, 
        "data_labels": {"value": True, "border": {}, "fill": {}, "font": {"color": "black"}},
    })
    
    chart.set_table()
    worksheet.insert_chart(chart_position, chart, {"x_offset": 25, "y_offset": 10})


@app.route('/')
def index():
    return '''
        <h1>Upload an Excel File</h1>
        <form action="/upload" method="post" enctype="multipart/form-data">
            <input type="file" name="file">
            <input type="submit" value="Upload">
        </form>
    '''

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return redirect(request.url)
    file = request.files['file']
    if file.filename == '':
        return redirect(request.url)
    if file and file.filename.endswith('.xlsx'):
        # Create a temporary directory to store the files
        temp_dir = tempfile.mkdtemp()

        # Process first output file: Data Tables
        wb = Workbook()
        wb.remove(wb.active)
        
        sheet_name = '強積⾦表現'
        source_data = pd.read_excel(file, sheet_name=sheet_name)
        extracted_data = source_data.iloc[0:4, [0, 14, 9]]
        extracted_data.columns = ['指數', '2024年6月回報（%)', '2024年1月回報（%)']
        extracted_data['2024年初⾄今回報（%)'] = extracted_data['2024年6月回報（%)'] - extracted_data['2024年1月回報（%)']
        final_data = extracted_data[['指數', '2024年6月回報（%)', '2024年初⾄今回報（%)']]
        ws = wb.create_sheet(title="強積⾦綜合指數6⽉升+1.1%2024年初⾄今升+5.5%")
        for col_num, column_title in enumerate(final_data.columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)
        for row_num, row in enumerate(final_data.itertuples(index=False), 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        create_table(ws, final_data, "強積⾦綜合指數6⽉升+1.1%2024年初⾄今升+5.5%")

        sheet_name = '2024年初至5月基金淨轉換額'
        source_data = pd.read_excel(file, sheet_name=sheet_name)
        extracted_data = source_data.iloc[0:26, [0, 1, 2]]
        ranking_column_name = '2024年(1月至5月)淨轉換額(百 萬 )'
        extracted_data['排名'] = extracted_data[ranking_column_name].rank(ascending=False, method='min').astype(int)
        final_data = extracted_data[['排名', '資產類別', ranking_column_name, '2024年(1月至5月) 淨轉換額%']]
        ws = wb.create_sheet(title="2024年初至5月基金淨轉換額")
        for col_num, column_title in enumerate(final_data.columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)
        for row_num, row in enumerate(final_data.itertuples(index=False), 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        create_table(ws, final_data, "2024年初至5月基金淨轉換額")

        sheet_name = '2024年初至5月基金淨轉換額'
        source_data = pd.read_excel(file, sheet_name=sheet_name)
        extracted_data = source_data.iloc[[26, 28], [0, 1, 3, 7]]
        extracted_data.columns = [' ', '2024年(1月至5月)淨轉換額(百 萬 )', '2023全年淨轉換額(百 萬 )', '2022全年淨轉換額(百 萬 )']
        ws = wb.create_sheet(title="歷年基⾦淨轉換額同比")
        for col_num, column_title in enumerate(extracted_data.columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)
        for row_num, row in enumerate(extracted_data.itertuples(index=False), 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        create_table(ws, extracted_data, "歷年基⾦淨轉換額同比")

        sheet_name = '2024年初至5月基金淨轉換額'
        source_data = pd.read_excel(file, sheet_name=sheet_name)
        extracted_data = source_data.iloc[[26, 27], [0, 1, 5, 9]]
        extracted_data.columns = [' ', '2024年(1月至5月)淨轉換額(百 萬 )', '2023年(1月至5月) 淨轉換額(百 萬 )', '2022年(1月至5月) 淨轉換額(百 萬 )']
        ws = wb.create_sheet(title="歷年基⾦淨轉換額同比(2)")
        for col_num, column_title in enumerate(extracted_data.columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)
        for row_num, row in enumerate(extracted_data.itertuples(index=False), 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        create_table(ws, extracted_data, "歷年基⾦淨轉換額同比(2)")

        sheet_name = '強積⾦表現'
        source_data = pd.read_excel(file, sheet_name=sheet_name)
        extracted_data = source_data.iloc[4:13, [0, 9, 14, 23]]
        extracted_data['2024年初至今'] = extracted_data['2024 Jun'] - extracted_data['2024 Jan']
        ranking_column_name = '2024年初至今'
        extracted_data['排名'] = extracted_data[ranking_column_name].rank(ascending=False, method='min').astype(int)
        final_data = extracted_data[['排名', '名稱', '2024年初至今', '2024 Jun', 'Market_Share ']]
        final_data.columns = ['排名', '固定收益基金指數成份基金種類', '2024年初至今', '6月回報 (%)', '總資產(%)(佔強積金市場所有資成份基金種類 產比例*)']
        ws = wb.create_sheet(title="2024年初⾄今強積⾦表現回顧(股票基金指數)")
        for col_num, column_title in enumerate(final_data.columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)
        for row_num, row in enumerate(final_data.itertuples(index=False), 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        create_table(ws, final_data, "2024年初⾄今強積⾦表現回顧(股票基金指數)")

        sheet_name = '強積⾦表現'
        source_data = pd.read_excel(file, sheet_name=sheet_name)
        extracted_data = source_data.iloc[22:30, [0, 9, 14, 23]]
        extracted_data['2024年初至今'] = extracted_data['2024 Jun'] - extracted_data['2024 Jan']
        ranking_column_name = '2024年初至今'
        extracted_data['排名'] = extracted_data[ranking_column_name].rank(ascending=False, method='min').astype(int)
        final_data = extracted_data[['排名', '名稱', '2024年初至今', '2024 Jun', 'Market_Share ']]
        final_data.columns = ['排名', '固定收益基金指數成份基金種類', '2024年初至今', '6月回報 (%)', '總資產(%)(佔強積金市場所有資成份基金種類 產比例*)']
        ws = wb.create_sheet(title="2024年初⾄今強積⾦表現回顧(固定收益基金指數)")
        for col_num, column_title in enumerate(final_data.columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)
        for row_num, row in enumerate(final_data.itertuples(index=False), 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        create_table(ws, final_data, "2024年初⾄今強積⾦表現回顧(固定收益基金指數)")

        sheet_name = '強積⾦表現'
        source_data = pd.read_excel(file, sheet_name=sheet_name)
        extracted_data = source_data.iloc[13:22, [0, 9, 14, 23]]
        extracted_data['2024年初至今'] = extracted_data['2024 Jun'] - extracted_data['2024 Jan']
        ranking_column_name = '2024年初至今'
        extracted_data['排名'] = extracted_data[ranking_column_name].rank(ascending=False, method='min').astype(int)
        final_data = extracted_data[['排名', '名稱', '2024年初至今', '2024 Jun', 'Market_Share ']]
        final_data.columns = ['排名', '股票基金指數成份基金種類', '2024年初至今', '6月回報 (%)', '總資產(%)(佔強積金市場所有資成份基金種類 產比例*)']
        ws = wb.create_sheet(title="2024年初⾄今強積⾦表現回顧(混合資產基金指數)")
        for col_num, column_title in enumerate(final_data.columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)
        for row_num, row in enumerate(final_data.itertuples(index=False), 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        create_table(ws, final_data, "2024年初⾄今強積⾦表現回顧(混合資產基金指數)")

        sheet_name = '強積⾦表現'  # Change this to the sheet name you want to read from
        source_data = pd.read_excel(file, sheet_name=sheet_name)
        extracted_data = source_data.iloc[[17, 18], [0, 2, 3, 4, 5, 6, 7, 8, 9, 14, 17]]
        extracted_data['2024年初⾄今'] = extracted_data['2024 Jun'] - extracted_data['2024 Jan']
        final_data = extracted_data[['名稱',           # Adjust or replace this with the correct column name
            '2017年',        # Adjust or replace this with the correct column name
            '2018年',        # Adjust or replace this with the correct column name
            '2019年',        # Adjust or replace this with the correct column name
            '2020年',        # Adjust or replace this with the correct column name
            '2021年',        # Adjust or replace this with the correct column name
            '2022年',        # Adjust or replace this with the correct column name
            '2023年',        # Adjust or replace this with the correct column name
            '2024年初⾄今',
            '推出⾄今累積回報'] ]
        final_data.columns = [' ',           # Adjust or replace this with the correct column name
            '2017年',        # Adjust or replace this with the correct column name
            '2018年',        # Adjust or replace this with the correct column name
            '2019年',        # Adjust or replace this with the correct column name
            '2020年',        # Adjust or replace this with the correct column name
            '2021年',        # Adjust or replace this with the correct column name
            '2022年',        # Adjust or replace this with the correct column name
            '2023年',        # Adjust or replace this with the correct column name
            '2024年初⾄今',
            '推出⾄今累積回報'] 
        ws = wb.create_sheet(title="預設投資策略(DIS)基⾦回報")
        for col_num, column_title in enumerate(final_data.columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)
        for row_num, row in enumerate(final_data.itertuples(index=False), 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        create_table(ws, final_data, "預設投資策略(DIS)基⾦回報")

        sheet_name = '強積⾦表現'  # Change this to the sheet name you want to read from
        source_data = pd.read_excel(file, sheet_name=sheet_name)
        extracted_data = source_data.iloc[[22, 16, 9, 4, 17], [0, 23, 20, 21, 19]]
        extracted_data.columns = [' ', '佔強積金市場所有資產比例', '年初至今最佳', '年初至今最差', '最佳與最差的基金回報差距']
        ws = wb.create_sheet(title="2024年初至今首五位最多人選擇的基金類別回報差距")
        for col_num, column_title in enumerate(extracted_data.columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)
        for row_num, row in enumerate(extracted_data.itertuples(index=False), 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        create_table(ws, extracted_data, "2024年初至今首五位最多人選擇的基金類別回報差距")

        sheet_name = '萬通信託下調管理費'  # Change this to the sheet name you want to read from
        source_data = pd.read_excel(file, sheet_name=sheet_name)
        extracted_data = source_data.iloc[0:14, 0:4]
        extracted_data.columns = [' ', '現時基金管理費', '由2024年9月26日起的基金管理費', '減幅']
        ws = wb.create_sheet(title="萬通信託下調管理費")
        for col_num, column_title in enumerate(extracted_data.columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)
        for row_num, row in enumerate(extracted_data.itertuples(index=False), 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        create_table(ws, extracted_data, "萬通信託下調管理費")

        sheet_name = '中國⼈壽下調管理費成分基金名稱'  # Change this to the sheet name you want to read from
        source_data = pd.read_excel(file, sheet_name=sheet_name)
        extracted_data = source_data.iloc[0:10, 0:4]
        ws = wb.create_sheet(title="中國⼈壽下調管理費")
        for col_num, column_title in enumerate(extracted_data.columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)
        for row_num, row in enumerate(extracted_data.itertuples(index=False), 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        create_table(ws, extracted_data, "中國⼈壽下調管理費")

        sheet_name = '強積⾦表現'  # Change this to the sheet name you want to read from
        source_data = pd.read_excel(file, sheet_name=sheet_name)
        extracted_data = source_data.iloc[0:4, [0, 18, 14, 9]]
        extracted_data.columns = ['指數', '指數值', '2024年6月回報（%)', '2024年1月回報（%)']
        extracted_data['2024年初⾄今回報（%)'] = extracted_data['2024年6月回報（%)'] - extracted_data['2024年1月回報（%)']
        final_data = extracted_data[['指數', '指數值', '2024年初⾄今回報（%)', '2024年6月回報（%)']]
        ws = wb.create_sheet(title="表⼀強積⾦市場整體表現及⼈均回報")
        for col_num, column_title in enumerate(final_data.columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)
        for row_num, row in enumerate(final_data.itertuples(index=False), 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        create_table(ws, final_data, "表⼀強積⾦市場整體表現及⼈均回報")

        sheet_name = '強積⾦表現'  # Change this to the sheet name you want to read from
        source_data = pd.read_excel(file, sheet_name=sheet_name)
        extracted_data = source_data.iloc[4:13, [0, 14, 9]]
        extracted_data['2024年初至今'] = extracted_data['2024 Jun'] - extracted_data['2024 Jan']
        ranking_column_name = '2024年初至今'
        extracted_data['排名'] = extracted_data[ranking_column_name].rank(ascending=False, method='min').astype(int)
        final_data = extracted_data[['排名', '名稱', '2024年初至今', '2024 Jun', ]]
        final_data.columns = ['排名', '股票基⾦指數', '2024 年初⾄今回報(%)', '6月回報 (%)']
        ws = wb.create_sheet(title="表⼆股票基⾦附屬指數表現排名(按2024年初⾄今回報排序)")
        for col_num, column_title in enumerate(final_data.columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)
        for row_num, row in enumerate(final_data.itertuples(index=False), 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        create_table(ws, final_data, "表⼆股票基⾦附屬指數表現排名(按2024年初⾄今回報排序)")

        sheet_name = '強積⾦表現'  # Change this to the sheet name you want to read from
        source_data = pd.read_excel(file, sheet_name=sheet_name)
        extracted_data = source_data.iloc[13:22, [0, 14, 9]]
        extracted_data['2024年初至今'] = extracted_data['2024 Jun'] - extracted_data['2024 Jan']
        ranking_column_name = '2024年初至今'
        extracted_data['排名'] = extracted_data[ranking_column_name].rank(ascending=False, method='min').astype(int)
        final_data = extracted_data[['排名', '名稱', '2024年初至今', '2024 Jun', ]]
        final_data.columns = ['排名', '混合資產基⾦指數 ', '2024 年初⾄今回報(%)', '6月回報 (%)']
        ws = wb.create_sheet(title="表三混合資產基⾦附屬指數表現排名(按2024年初⾄今回報排序)")
        for col_num, column_title in enumerate(final_data.columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)
        for row_num, row in enumerate(final_data.itertuples(index=False), 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        create_table(ws, final_data, "表三混合資產基⾦附屬指數表現排名(按2024年初⾄今回報排序)")

        sheet_name = '強積⾦表現'  # Change this to the sheet name you want to read from
        source_data = pd.read_excel(file, sheet_name=sheet_name)
        extracted_data = source_data.iloc[22:30, [0, 14, 9]]
        extracted_data['2024年初至今'] = extracted_data['2024 Jun'] - extracted_data['2024 Jan']
        ranking_column_name = '2024年初至今'
        extracted_data['排名'] = extracted_data[ranking_column_name].rank(ascending=False, method='min').astype(int)
        final_data = extracted_data[['排名', '名稱', '2024年初至今', '2024 Jun']]
        final_data.columns = ['排名', '固定收益基⾦指數', '2024 年初⾄今回報(%)', '6月回報 (%)']
        ws = wb.create_sheet(title="表四固定收益基⾦附屬指數表現排名(按2024年初⾄今回報排序)")
        for col_num, column_title in enumerate(final_data.columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)
        for row_num, row in enumerate(final_data.itertuples(index=False), 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        create_table(ws, final_data, "表四固定收益基⾦附屬指數表現排名(按2024年初⾄今回報排序)")

        sheet_name = '2024年初至5月基金淨轉換額'  # Change this to the sheet name you want to read from
        source_data = pd.read_excel(file, sheet_name=sheet_name)
        extracted_data = source_data.iloc[[26], [0, 1, 5, 9]]
        extracted_data.columns = [' ', '2024年(1月至5月)', '2023年(1月至5月)', '2022年(1月至5月)']
        ws = wb.create_sheet(title="表五歷年基⾦淨轉換額同比")
        for col_num, column_title in enumerate(extracted_data.columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)
        for row_num, row in enumerate(extracted_data.itertuples(index=False), 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        create_table(ws, extracted_data, "表五歷年基⾦淨轉換額同比")

        sheet_name = '2024年初至5月基金淨轉換額'  # Change this to the sheet name you want to read from
        source_data = pd.read_excel(file, sheet_name=sheet_name)
        extracted_data = source_data.iloc[[5, 13, 18, 6, 4], [0, 1]]
        extracted_data.columns = ['基⾦類別', '年初⾄今資⾦淨轉入(百萬港元)']
        ws = wb.create_sheet(title="表七2024年初⾄今淨轉入轉出資產類別(頭5名)")
        for col_num, column_title in enumerate(extracted_data.columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)
        for row_num, row in enumerate(extracted_data.itertuples(index=False), 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        create_table(ws, extracted_data, "表七2024年初⾄今淨轉入轉出資產類別(頭5名)")

        sheet_name = '2024年初至5月基金淨轉換額'  # Change this to the sheet name you want to read from
        source_data = pd.read_excel(file, sheet_name=sheet_name)
        extracted_data = source_data.iloc[[0, 12, 1, 2, 11], [0, 1]]
        extracted_data.columns = ['基⾦類別', '年初⾄今資⾦淨轉出(百萬港元)']
        ws = wb.create_sheet(title="表七2024年初⾄今淨轉入轉出資產類別(頭5名)(2)")
        for col_num, column_title in enumerate(extracted_data.columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)
        for row_num, row in enumerate(extracted_data.itertuples(index=False), 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        create_table(ws, extracted_data, "表七2024年初⾄今淨轉入轉出資產類別(頭5名)(2)")

        # Save the first workbook to a file
        output1_path = os.path.join(temp_dir, 'data_tables.xlsx')
        wb.save(output1_path)
        
        # Process second output file: Charts
        writer = pd.ExcelWriter(os.path.join(temp_dir, "pandas_chart_combined.xlsx"), engine="xlsxwriter")

        for sheet_name in pd.ExcelFile(file).sheet_names:
            df = pd.read_excel(file, sheet_name=sheet_name)
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        workbook = writer.book
        worksheet1 = writer.sheets['強積金綜合指數']
        create_combined_chart(
            workbook=workbook,
            worksheet=worksheet1,
            col_series_range="=強積金綜合指數!$B$2:$B$8",
            line_series_range=(['強積金綜合指數', 1, 0, 8, 0], ['強積金綜合指數', 1, 2, 8, 2]),
            chart_position="E2",
            chart_title="強積金綜合指數"
        )

        worksheet2 = writer.sheets['2024年初至5月強積金資產類別淨轉換估算(以十億港元計)']
        create_column_chart(
            workbook=workbook,
            worksheet=worksheet2,
            categories_range=['2024年初至5月強積金資產類別淨轉換估算(以十億港元計)', 1, 0, 3, 0],
            values_range="=2024年初至5月強積金資產類別淨轉換估算(以十億港元計)!$H$2:$H$4",
            chart_position="D5",
            chart_title="2024年初至5月強積金資產類別淨轉換估算(以十億港元計)"
        )

        worksheet3 = writer.sheets['強積⾦表現']
        create_custom_chart(
            workbook=workbook,
            worksheet=worksheet3,
            chart_position="D2"
        )

        chart1 = workbook.add_chart({"type": "column"})
        chart1.add_series({
            "name": "=強積⾦表現!$U$1",
            "categories": "=強積⾦表現!$A$24:$A$31",
            "values": "=強積⾦表現!$U$24:$U$31",
            "fill": {"color": "yellow"}, 
            "data_labels": {"value": True, "border": {}, "fill": {}, "font":{"color": "green"}},
        })
        chart1.add_series({
            "name": ["強積⾦表現", 0, 21],
            "categories": ["強積⾦表現", 23, 0, 30, 0],
            "values": ["強積⾦表現", 23, 21, 30, 21],
            "fill": {"color": "yellow"}, 
            "data_labels": {"value": True, "border": {}, "fill": {},"font":{"color": "red"}},
        })
        chart1.add_series({
            "name": ["強積⾦表現", 0, 22],
            "categories": ["強積⾦表現", 23, 0, 30, 0],
            "values": ["強積⾦表現", 23, 22, 30, 22],
            "fill": {"color": "black"}, 
            "data_labels": {"value": True, "border": {}, "fill": {},"font":{"color": "black"}},
        })
        chart1.set_table()
        worksheet3.insert_chart("H2", chart1, {"x_offset": 25, "y_offset": 10})

        chart2 = workbook.add_chart({"type": "column"})
        chart2.add_series({
            "name": "=強積⾦表現!$U$1",
            "categories": "=強積⾦表現!$A$15:$A$23",
            "values": "=強積⾦表現!$U$15:$U$23",
            "fill": {"color": "yellow"}, 
            "data_labels": {"value": True, "border": {}, "fill": {}, "font":{"color": "green"}},
        })
        chart2.add_series({
            "name": ["強積⾦表現", 0, 21],
            "categories": ["強積⾦表現", 14, 0, 22, 0],
            "values": ["強積⾦表現", 14, 21, 22, 21],
            "fill": {"color": "yellow"}, 
            "data_labels": {"value": True, "border": {}, "fill": {},"font":{"color": "red"}},
        })
        chart2.add_series({
            "name": ["強積⾦表現", 0, 22],
            "categories": ["強積⾦表現", 14, 0, 22, 0],
            "values": ["強積⾦表現", 14, 22, 22, 22],
            "fill": {"color": "black"}, 
            "data_labels": {"value": True, "border": {}, "fill": {},"font":{"color": "black"}},
        })
        chart2.set_table()
        worksheet3.insert_chart("L2", chart2, {"x_offset": 25, "y_offset": 10})


        worksheet4 = writer.sheets['強積⾦表現']
        chart2 = workbook.add_chart({"type": "bar"})

        chart2.add_series({
            "name": '=強積⾦表現!$A$24',
            "categories": '=強積⾦表現!$A$24',
            "values": '=強積⾦表現!$T$24',
            "data_labels": {"value": True, "border": {}, "fill": {}},
        })

        chart2.add_series({
            "name": '=強積⾦表現!$A$18',
            "categories": '=強積⾦表現!$A$18',
            "values": '=強積⾦表現!$T$18',
            "data_labels": {"value": True, "border": {}, "fill": {}},
        })

        chart2.add_series({
            "name": '=強積⾦表現!$A$11',
            "categories": '=強積⾦表現!$A$11',
            "values": '=強積⾦表現!$T$11',
            "data_labels": {"value": True, "border": {}, "fill": {}},
        })

        chart2.add_series({
            "name": '=強積⾦表現!$A$7',
            "categories": '=強積⾦表現!$A$7',
            "values": '=強積⾦表現!$T$7',
            "data_labels": {"value": True, "border": {}, "fill": {}},
        })

        chart2.add_series({
            "name": '=強積⾦表現!$A$19',
            "categories": '=強積⾦表現!$A$19',
            "values": '=強積⾦表現!$T$19',
            "data_labels": {"value": True, "border": {}, "fill": {}},
        })

        chart2.set_title({"name": "2024年初至今最佳與最差的基金回報差距"})
        chart2.set_x_axis({
            'position_axis': 'on_tick',
            'label_position': 'low'
        })

        chart2.set_y_axis({
            'name': '',
            'label_position': 'low'
        })
        # Chart 7: column Chart with Multiple Series for "output_table
        worksheet4.insert_chart("D25", chart2, {"x_offset": 25, "y_offset": 10})

        worksheet5 = writer.sheets['強積⾦表現']
        chart5 = workbook.add_chart({"type": "column"})

        chart5.add_series({
            "name": "=強積⾦表現!$A$24",
            "categories": "=強積⾦表現!$C$1:$O$1",
            "values": "=強積⾦表現!$C$24:$O$24",
            "data_labels": {"value": True, "border": {}, "fill": {}},
        })

        chart5.add_series({
            "name": ["強積⾦表現", 28, 0],
            "categories": ["強積⾦表現", 0, 2, 0, 14],
            "values": ["強積⾦表現", 28, 2, 28, 14],
            "data_labels": {"value": True, "border": {}, "fill": {}},
        })

        chart5.set_title({"name": "強積⾦保守基⾦ vs 環球債券基⾦"})
        chart5.set_x_axis({"label_position": "low"})

        worksheet5.insert_chart("D2", chart5, {"x_offset": 25,"y_offset": 10})
        writer.close()

        output2_path = os.path.join(temp_dir, "pandas_chart_combined.xlsx")
        
        # Provide download links for both files
        return render_template_string('''
            <h1>Files Generated</h1>
            <a href="{{ url_for('download_file', path=output1_path) }}">Download Data Tables</a><br>
            <a href="{{ url_for('download_file', path=output2_path) }}">Download Charts</a>
        ''', output1_path=output1_path, output2_path=output2_path)

    return redirect(request.url)

@app.route('/download')
def download_file():
    file_path = request.args.get('path')
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        return "File not found", 404

if __name__ == '__main__':
    app.run(debug=True)
